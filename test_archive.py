import os
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
import xlrd
from path import resources_path, zip_path
from pypdf import PdfReader


def test_archive():
    resources_files = os.listdir(resources_path)
    with ZipFile(zip_path, 'r') as zf:
        assert resources_files == ZipFile.namelist(zf)


def test_pdf():
    with ZipFile(zip_path, 'r') as zf:
        original_file = PdfReader(os.path.join(resources_path, 'Python Testing with Pytest (Brian Okken).pdf'))
        archived_file = PdfReader(zf.open('Python Testing with Pytest (Brian Okken).pdf'))
        assert len(original_file.pages) == len(archived_file.pages)
        assert original_file.pages[3].extract_text() == archived_file.pages[3].extract_text()


def test_txt():
    with ZipFile(zip_path, 'r') as zf:
        with open(os.path.join(resources_path, 'Hello.txt')) as txt:
            original_file = txt.read()
            assert original_file == zf.read('Hello.txt').decode()


def test_xls():
    with ZipFile(zip_path, 'r') as zf:
        with zf.open('file_example_XLS_10.xls') as xls:
            temp = xls.read()
            archived_file = xlrd.open_workbook(file_contents=temp)
        original_file = xlrd.open_workbook(os.path.join(resources_path, 'file_example_XLS_10.xls'))
        assert original_file.nsheets == archived_file.nsheets
        assert original_file.sheet_names() == archived_file.sheet_names()
        sheet_original = original_file.sheet_by_index(0)
        sheet_archived = archived_file.sheet_by_index(0)
        assert sheet_original.ncols == sheet_archived.ncols
        assert sheet_original.nrows == sheet_archived.nrows


def test_xlsx():
    with ZipFile(zip_path, 'r') as zf:
        archived_file = load_workbook(zf.open('file_example_XLSX_50.xlsx'))
        original_file = load_workbook(os.path.join(resources_path, 'file_example_XLSX_50.xlsx'))
        assert original_file.sheetnames == archived_file.sheetnames
        sheet_original = original_file.active
        sheet_archived = archived_file.active
        row_count = sheet_original.max_row + 1
        column_count = sheet_original.max_column + 1
        print(row_count, column_count)
        for i in range(1, row_count):
            for j in range(1, column_count):
                assert sheet_original.cell(row=i, column=j).value == sheet_archived.cell(row=i, column=j).value










